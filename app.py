from flask import Flask, render_template, request, jsonify, redirect, url_for, flash
from openpyxl import load_workbook
import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'  # セッション用の秘密鍵

EXCEL_FILES = [
    {
        'filename': '小テスト Retrieved from コーパス4500 4th Edition.xlsx',
        'display_name': 'コーパス4500',
        'category': '英単語',
        'word_column': 0,
        'meaning_column': 1,
        'source_label': '英単語',
        'default_title': '英単語テスト',
    },
    {
        'filename': 'ターゲット1900.xlsx',
        'display_name': 'ターゲット1900',
        'category': '英単語',
        'word_column': 0,
        'meaning_column': 1,
        'source_label': '英単語',
        'default_title': '英単語テスト',
    },
    {
        'filename': 'システム英単語.xlsx',
        'display_name': 'システム英単語',
        'category': '英単語',
        'word_column': 0,
        'meaning_column': 1,
        'source_label': '英単語',
        'default_title': '英単語テスト',
    },
    {
        'filename': 'LEAP.xlsx',
        'display_name': 'LEAP',
        'category': '英単語',
        'word_column': 0,
        'meaning_column': 1,
        'source_label': '英単語',
        'default_title': '英単語テスト',
    },
    {
        'filename': '英検2級_出る順パス単　.xlsx',
        'display_name': '英検2級 出る順パス単',
        'category': '英単語',
        'word_column': 0,
        'meaning_column': 1,
        'source_label': '英単語',
        'default_title': '英単語テスト',
    },
    {
        'filename': '英検準1級_出る順パス単.xlsx',
        'display_name': '英検準1級 出る順パス単',
        'category': '英単語',
        'word_column': 0,
        'meaning_column': 1,
        'source_label': '英単語',
        'default_title': '英単語テスト',
    },
    {
        'filename': '単語帳EX 準1級.xlsx',
        'display_name': '単語帳EX 準1級',
        'category': '英単語',
        'word_column': 0,
        'meaning_column': 1,
        'source_label': '英単語',
        'default_title': '英単語テスト',
    },
    {
        'filename': '速読英熟語.xlsx',
        'display_name': '速読英熟語',
        'category': '英熟語',
        'word_column': 0,
        'meaning_column': 1,
        'source_label': '英熟語',
        'default_title': '英熟語テスト',
    },
    {
        'filename': '古文単語/古文単語315.xlsx',
        'display_name': '古文単語315',
        'category': '古文単語',
        'word_column': 1,
        'meaning_column': 2,
        'source_label': '古文単語',
        'default_title': '古文単語テスト',
    },
    {
        'filename': '出る順に学ぶ頻出古文単語400.xlsx',
        'display_name': '出る順に学ぶ頻出古文単語400',
        'category': '古文単語',
        'word_column': 0,
        'meaning_column': 1,
        'source_label': '古文単語',
        'default_title': '古文単語テスト',
    },
]

HEADER_VALUES = {
    '単語名', '単語', '英単語', '古文単語', 'word', 'Word', 'WORD',
    '意味', 'meaning', 'Meaning', 'MEANING',
}


def get_file_config(filename=None):
    """ファイルごとの読み込み設定を取得する"""
    if filename is None:
        return EXCEL_FILES[0]

    for config in EXCEL_FILES:
        if config['filename'] == filename:
            return config

    return {
        'filename': filename,
        'display_name': os.path.basename(filename),
        'category': '単語帳',
        'word_column': 0,
        'meaning_column': 1,
        'source_label': '単語',
        'default_title': '小テスト',
    }


def load_excel_data(filename=None):
    """Excelファイルからデータを読み込む"""
    try:
        config = get_file_config(filename)
        filename = config['filename']
        word_column = config['word_column']
        meaning_column = config['meaning_column']
        
        # Excelファイルを読み込み
        workbook = load_workbook(filename)
        worksheet = workbook.active
        
        data = []
        row_num = 1
        
        for row in worksheet.iter_rows(min_row=1, values_only=True):
            if len(row) <= max(word_column, meaning_column):
                continue

            word = row[word_column]
            meaning = row[meaning_column]

            if word is None or meaning is None:
                continue

            word = str(word).strip()
            meaning = str(meaning).strip()

            if not word or not meaning or word in HEADER_VALUES or meaning in HEADER_VALUES:
                continue

            data.append({
                'row_number': row_num,
                'word': word,
                'meaning': meaning
            })
            row_num += 1
        
        return data
    except Exception as e:
        print(f"Excelファイルの読み込みエラー: {e}")
        return []

def get_available_files():
    """利用可能なExcelファイルのリストを取得"""
    files = []

    for config in EXCEL_FILES:
        file = config['filename']
        if os.path.exists(file):
            files.append(config)
        else:
            # ファイルが見つからない場合は、サンプルファイルの存在を確認
            sample_file = f'sample_data/{file.replace(".xlsx", "_sample.xlsx")}'
            if os.path.exists(sample_file):
                sample_config = config.copy()
                sample_config['filename'] = sample_file
                sample_config['display_name'] = config['display_name'] + ' (サンプル)'
                files.append(sample_config)
    
    return files

@app.route('/')
def index():
    """メインページ"""
    data = load_excel_data()
    available_files = get_available_files()
    available_categories = []
    for file in available_files:
        if file['category'] not in available_categories:
            available_categories.append(file['category'])
    return render_template(
        'index.html',
        data=data,
        available_files=available_files,
        available_categories=available_categories
    )


@app.route('/generate_test', methods=['POST'])
def generate_test():
    """指定された行番号のテストを生成"""
    try:
        data = request.get_json()
        selected_rows = data.get('selected_rows', [])
        
        if not selected_rows:
            return jsonify({'error': '行番号が選択されていません'}), 400
        
        # 選択されたファイルのデータを読み込み
        selected_file = data.get('selected_file')
        if selected_file:
            excel_data = load_excel_data(selected_file)
        else:
            excel_data = load_excel_data()
        
        # 選択された行番号のデータを取得
        test_data = []
        print(f"選択された行番号: {selected_rows}")
        print(f"利用可能なデータ数: {len(excel_data)}")
        
        for row_num in selected_rows:
            found = False
            for item in excel_data:
                if item['row_number'] == row_num:
                    test_data.append(item)
                    print(f"行番号 {row_num}: {item['word']} - {item['meaning']}")
                    found = True
                    break
            if not found:
                print(f"警告: 行番号 {row_num} のデータが見つかりません")
        
        print(f"最終的なテストデータ数: {len(test_data)}")
        
        if not test_data:
            return jsonify({'error': '選択された行番号のデータが見つかりません'}), 400
        
        return jsonify({
            'success': True,
            'test_data': test_data
        })
        
    except Exception as e:
        return jsonify({'error': f'エラーが発生しました: {str(e)}'}), 500

@app.route('/get_all_data')
def get_all_data():
    """全データを取得"""
    data = load_excel_data()
    return jsonify(data)

@app.route('/load_file', methods=['POST'])
def load_file():
    """指定されたファイルのデータを読み込み"""
    try:
        data = request.get_json()
        filename = data.get('filename')
        
        if not filename:
            return jsonify({'error': 'ファイル名が指定されていません'}), 400
        
        file_data = load_excel_data(filename)
        
        if not file_data:
            return jsonify({'error': 'ファイルの読み込みに失敗しました'}), 400
        
        return jsonify({
            'success': True,
            'data': file_data
        })
        
    except Exception as e:
        return jsonify({'error': f'エラーが発生しました: {str(e)}'}), 500

@app.route('/get_available_files')
def get_available_files_api():
    """利用可能なファイル一覧を取得"""
    files = get_available_files()
    return jsonify(files)

@app.route('/contact')
def contact():
    """問い合わせフォームページ"""
    return render_template('contact.html')

@app.route('/submit_contact', methods=['POST'])
def submit_contact():
    """問い合わせフォームの送信処理"""
    try:
        name = request.form.get('name', '').strip()
        subject = request.form.get('subject', '').strip()
        message = request.form.get('message', '').strip()
        
        # バリデーション
        if not name or not subject or not message:
            flash('すべての項目を入力してください。', 'error')
            return redirect(url_for('contact'))
        
        # メール送信（実際の送信は環境変数で設定）
        email_sent = send_contact_email(name, subject, message)
        
        if email_sent:
            flash('お問い合わせありがとうございます。メールを送信しました。内容を確認の上、ご連絡いたします。', 'success')
        else:
            flash('お問い合わせを受け付けましたが、メール送信に失敗しました。しばらくしてから再度お試しください。', 'warning')
        
        return redirect(url_for('contact'))
        
    except Exception as e:
        flash(f'エラーが発生しました: {str(e)}', 'error')
        return redirect(url_for('contact'))

def send_contact_email(name, subject, message):
    """問い合わせメールを送信"""
    try:
        # メール設定（環境変数から取得、なければデフォルト値）
        smtp_server = os.environ.get('SMTP_SERVER', 'smtp.gmail.com')
        smtp_port = int(os.environ.get('SMTP_PORT', '587'))
        smtp_username = os.environ.get('SMTP_USERNAME', '')
        smtp_password = os.environ.get('SMTP_PASSWORD', '')
        contact_email = os.environ.get('CONTACT_EMAIL', 'contact@example.com')
        
        # デバッグ情報を出力
        print(f"=== メール送信デバッグ情報 ===")
        print(f"SMTP_SERVER: {smtp_server}")
        print(f"SMTP_PORT: {smtp_port}")
        print(f"SMTP_USERNAME: {'設定済み' if smtp_username else '未設定'}")
        print(f"SMTP_PASSWORD: {'設定済み' if smtp_password else '未設定'}")
        print(f"CONTACT_EMAIL: {contact_email}")
        print(f"===============================")
        
        # メール内容
        msg = MIMEMultipart()
        msg['From'] = smtp_username
        msg['To'] = contact_email
        msg['Subject'] = f'[テスト問題ジェネレーター] {subject}'
        
        body = f"""
新しい英単語帳の追加要望が届きました。

送信者名: {name}
件名: {subject}

メッセージ:
{message}

---
このメールはテスト問題ジェネレーターの問い合わせフォームから送信されました。
        """
        
        msg.attach(MIMEText(body, 'plain', 'utf-8'))
        
        # メール送信（SMTP設定がある場合のみ）
        if smtp_username and smtp_password:
            print("SMTP設定が検出されました。メール送信を試行します...")
            server = smtplib.SMTP(smtp_server, smtp_port)
            server.starttls()
            server.login(smtp_username, smtp_password)
            text = msg.as_string()
            server.sendmail(smtp_username, contact_email, text)
            server.quit()
            print("メール送信が完了しました。")
            return True  # 送信成功を返す
        else:
            # SMTP設定がない場合はログに出力
            print("SMTP設定が不完全です。メール送信をスキップします。")
            print(f"問い合わせメール（送信設定なし）:")
            print(f"送信者: {name}")
            print(f"件名: {subject}")
            print(f"メッセージ: {message}")
            return False  # 送信失敗を返す
            
    except Exception as e:
        print(f"メール送信エラー: {e}")
        print(f"エラーの詳細: {type(e).__name__}")
        return False  # エラー時は送信失敗を返す

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5001))
    app.run(debug=False, host='0.0.0.0', port=port)

# Gunicorn用の設定
if __name__ != '__main__':
    # 本番環境用の設定
    pass
